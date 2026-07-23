#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Factor Analysis -- 시가총액 & 거래대금 분기별 성장률 분석

유니버스 : 현재 시총 상위 Top N (고정 유니버스)
Factor 1 : 시총 성장률 (QoQ %)
Factor 2 : 거래대금 증가율 (QoQ %)

기간  : 최근 4년 (16분기)
가중치: 선형 -- 1.00 (4년전 Q1) -> 1.75 (최근 Q16), 분기당 +0.05

실행:
  source /data/tibero7/t7.profile
  source /data/frame/.venv/bin/activate
  python3 factor_analysis.py          # 미장 (기본)
  python3 factor_analysis.py --kr     # 국장
  python3 factor_analysis.py --years 4 --top 100
  python3 factor_analysis.py --refresh        # ETL로 데이터 최신화 후 분석
"""

from __future__ import annotations

import sys
import os
import argparse
import subprocess
import glob
import time
from datetime import date

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import jaydebeapi
except ImportError:
    print("pip install jaydebeapi JPype1")
    sys.exit(1)

# DB 설정
TIBERO_HOST = "localhost"
TIBERO_PORT = 44123
TIBERO_SID  = "tibero"
TIBERO_USER = os.environ.get("TIBERO_USER", "sys")
TIBERO_PASS = os.environ.get("TIBERO_PASS", "")
JDBC_JAR    = "/data/tibero7/tibero7/client/lib/jar/tibero7-jdbc.jar"
JDBC_CLASS  = "com.tmax.tibero.jdbc.TbDriver"

# 파라미터
YEARS = 4
TOP_N = 100

# 선형 가중치
W_START = 1.00
W_STEP  = 0.05

# 종합 점수 가중치 — IC 검증(33분기) 기반 재설계 (2026-06)
#   * 검증서 |t|>2(유의)한 팩터는 하나도 없었고, 매출(SEC) 계열만 전·후반 부호가 안정적
#     이라 거기로 '가볍게' 틸트(과적합 경계해 정밀 가중은 지양).
#   * IC≈0 이거나 부호가 뒤집힌 모멘텀(시총가속/거래대금증가/거래대금일관)은 점수에서 제거.
#   * 과거 일자별 데이터가 없어 검증 불가한 PEG/ROE/부채비율은 '점수'가 아니라
#     '게이트(통과/탈락 필터)'로 강등 → GATES 참고. (NaN-aware 정규화라 합계=1.0 불필요)
COMPOSITE_WEIGHTS = {
    # ── 매출(SEC) — IC 검증서 유일하게 부호가 살아남은 축 ──
    "rev_acceleration": 0.20,   # 매출 YoY 가속도 (코어; 전·후반 둘 다 +IC)
    "revenue_growth":   0.12,   # 매출 성장률 TTM (서브코어)
    "rev_consistency":  0.06,   # 매출 YoY 일관성 (후반 소멸 → 감비중·관찰)
    # ── 품질(약) ──
    "operating_margin": 0.03,   # 영업이익률 (부호 불안정 → 최소)
    # ── 잔여 모멘텀 — IC≈0 이라 흔적만 ──
    "시총_성장률":       0.02,
    "시총_일관성":       0.02,
    "거래대금_가속도":    0.01,
    # ── 점수 제외(0): 시총_가속도(부호반전)·거래대금_증가율(음IC)·거래대금_일관성(≈0)
    # ── 게이트로 강등: peg / roe / debt_to_equity  (아래 GATES 참고)
}
INVERTED_FACTORS = set()        # 점수에 들어가는 팩터는 전부 '높을수록 좋음'

# 검증불가 팩터(과거 일자별 데이터 없음) → 점수 가중에서 빼고 '게이트(필터)'로만 사용.
#   값이 분명히 나쁜 종목만 후보에서 제외(값이 없으면 통과 — 결측으로 탈락시키지 않음).
USE_GATES    = True
GATE_ROE_MIN = 0.0      # ROE 음수(자본 대비 적자) 제외
GATE_DE_MAX  = 500.0    # 부채비율(D/E) 과다 제외

# 안전(참고지표) — IC 검증(41분기) 결과 '안전↔저수익' 트레이드오프라 점수에 합산하지 않음.
#   변동성/하방편차는 |t|<1 로 비유의하지만 부호가 안정적(고변동=고수익) → 게이트도 위험(승자 절단).
#   따라서 '안전도/변동성/MDD/사분면'을 점수와 분리된 참고 컬럼으로만 표기(사용자가 trade-off 판단).
SAFETY_WIN      = 252   # 일별 trailing 거래일(약 1년)
SAFETY_MINVALID = 200   # 윈도우 내 최소 유효 일별수익률(미만이면 안전도 NaN)
PEG_CAP          = 50.0    # PEG 상한
DE_CAP           = 500.0   # D/E 상한 (아웃라이어 처리)
REV_ACCEL_CAP    = 50.0    # rev_acceleration ±cap (극단값 왜곡 방지)
MIN_REV_QUARTERS = 12      # 최소 분기 수 미만 → rev 팩터 NaN (신생·분사 종목 필터)
EXCLUDE_CODES    = {"GOOG"} # 중복 주식 클래스 제거 (GOOGL 유지)

# 티어 기준 (현재 시총 순위 기반)
TIER_BINS  = [0, 10, 30, 50, 100]
TIER_NAMES = ["Tier1 (1-10)", "Tier2 (11-30)", "Tier3 (31-50)", "Tier4 (51-100)"]


def get_conn():
    url = f"jdbc:tibero:thin:@{TIBERO_HOST}:{TIBERO_PORT}:{TIBERO_SID}"
    return jaydebeapi.connect(JDBC_CLASS, url, [TIBERO_USER, TIBERO_PASS], JDBC_JAR)


# ── STEP 0: 현재 유니버스 (고정) ──────────────────────────────────────────────

def get_current_universe(conn, top_n=100, table="daily_marcap_us"):
    """
    가장 최근 거래일 기준 시총 상위 top_n 종목 코드 + 현재 순위 반환
    이 유니버스를 고정해서 모든 과거 데이터를 조회함
    """
    sql = f"""
    SELECT code, rank
    FROM {table}
    WHERE date_ = (SELECT MAX(date_) FROM {table})
      AND rank <= {top_n + 10}
    ORDER BY rank
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()
    df = pd.DataFrame(rows, columns=["code", "current_rank"])
    df["current_rank"] = df["current_rank"].astype(int)
    return df


def get_name_map(conn, table="ticker_master"):
    """code -> name 매핑 (ticker_master / ticker_master_us)."""
    cur = conn.cursor()
    cur.execute(f"SELECT code, name FROM {table}")
    rows = cur.fetchall()
    cur.close()
    return {code: name for code, name in rows}


def assign_tier(rank):
    for i in range(len(TIER_BINS) - 1):
        if TIER_BINS[i] < rank <= TIER_BINS[i + 1]:
            return TIER_NAMES[i]
    return "기타"


# ── STEP 1: 분기말 날짜 ───────────────────────────────────────────────────────

def get_quarter_end_dates(conn, years=4, table="daily_marcap_us"):
    sql = f"""
    SELECT MAX(date_) AS qdate,
           EXTRACT(YEAR  FROM date_) AS yr,
           CEIL(EXTRACT(MONTH FROM date_) / 3) AS q_num
    FROM {table}
    WHERE date_ >= ADD_MONTHS(CURRENT_DATE, -{years * 12})
    GROUP BY EXTRACT(YEAR FROM date_), CEIL(EXTRACT(MONTH FROM date_) / 3)
    ORDER BY qdate
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()
    out = [(r[0], int(r[1]), int(r[2])) for r in rows]
    # 진행 중인 분기 제외: 마지막 qdate가 달력상 분기말보다 4일 넘게 이르면 미완성 분기
    # (포함 시 부분 합계·짧은 기간 변동이 완성 분기 QoQ처럼 최대 가중치를 받는 왜곡 발생)
    if out:
        qd, yr, qn = out[-1]
        q_end = pd.Timestamp(year=yr, month=qn * 3, day=1) + pd.offsets.MonthEnd(0)
        if (q_end - pd.Timestamp(str(qd)[:10])).days > 4:
            print(f"  진행 중 분기 제외: {yr} Q{qn} (마지막 거래일 {str(qd)[:10]} < 분기말 {q_end.date()})")
            out = out[:-1]
    return out


# ── STEP 2-A: Factor 1 시가총액 ───────────────────────────────────────────────

def get_quarterly_marcap(conn, qdates, universe_codes, table="daily_marcap_us"):
    """
    고정 유니버스 종목들의 분기말 시가총액 조회
    (rank 필터 없이 code IN 으로 조회 → 과거에 순위 밖이어도 가져옴)
    """
    if not qdates or not universe_codes:
        return pd.DataFrame()

    in_dates = ", ".join(f"TO_DATE('{str(r[0])[:10]}', 'YYYY-MM-DD')" for r in qdates)
    in_codes = ", ".join(f"'{c}'" for c in universe_codes)

    sql = f"""
    SELECT TO_CHAR(m.date_, 'YYYY-MM-DD') AS date_str,
           EXTRACT(YEAR  FROM m.date_) AS yr,
           CEIL(EXTRACT(MONTH FROM m.date_) / 3) AS q_num,
           m.code,
           m.marcap
    FROM {table} m
    WHERE m.date_ IN ({in_dates})
      AND m.code  IN ({in_codes})
    ORDER BY m.date_, m.code
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()
    df = pd.DataFrame(rows, columns=["date_str", "yr", "q_num", "code", "marcap"])
    df["marcap"] = pd.to_numeric(df["marcap"], errors="coerce")
    return df


# ── STEP 2-B: Factor 2 거래대금 ───────────────────────────────────────────────

def get_quarterly_amount(conn, qdates, universe_codes,
                         price_table="daily_price_us"):
    """
    고정 유니버스 종목들의 분기별 거래대금 합계 조회
    분기 시작일: Q1=1월, Q2=4월, Q3=7월, Q4=10월 1일
    """
    if not qdates or not universe_codes:
        return pd.DataFrame()

    in_codes = ", ".join(f"'{c}'" for c in universe_codes)
    frames = []

    for qdate, yr, q_num in qdates:
        qdate_str   = str(qdate)[:10]
        q_start_mo  = (q_num - 1) * 3 + 1
        q_start_str = f"{yr}-{q_start_mo:02d}-01"

        sql = f"""
        SELECT code, SUM(amount) AS total_amount
        FROM {price_table}
        WHERE date_ >= TO_DATE('{q_start_str}', 'YYYY-MM-DD')
          AND date_ <= TO_DATE('{qdate_str}',   'YYYY-MM-DD')
          AND code  IN ({in_codes})
        GROUP BY code
        """
        cur = conn.cursor()
        cur.execute(sql)
        for r in cur.fetchall():
            frames.append({
                "date_str":     qdate_str,
                "yr":           yr,
                "q_num":        q_num,
                "code":         r[0],
                "total_amount": float(r[1]) if r[1] else None,
            })
        cur.close()
        print(f"  거래대금 조회: {yr} Q{q_num} ({qdate_str})")

    return pd.DataFrame(frames)


# ── STEP 2-C: 매출 YoY 일관성 / 가속도 (SEC EDGAR) ───────────────────────────

def get_revenue_factors(conn, universe_codes, yoy_quarters=8, table="quarterly_financials_us"):
    """
    quarterly_financials_us/kr 에서 분기 매출 히스토리를 읽어
    YoY 일관성 / 가속도를 계산해 반환.

    - yoy_quarters : YoY 계산에 사용할 최근 분기 수 (기본 8 = 2년)
      → YoY를 계산하려면 1년 전 같은 분기 필요 → 실제로 yoy_quarters+4 이상 있어야 함
    - rev_consistency  = YoY > 0 인 분기 수 / 전체 YoY 분기 수  (0~1)
    - rev_acceleration = 최근 4Q 평균 YoY − 전체 평균 YoY
    """
    if not universe_codes:
        return pd.DataFrame(columns=["code", "rev_consistency", "rev_acceleration",
                                     "rev_quarters"])

    in_codes = ", ".join(f"'{c}'" for c in universe_codes)
    sql = f"""
    SELECT code, TO_CHAR(end_date, 'YYYY-MM-DD') AS end_date, revenue
    FROM {table}
    WHERE code IN ({in_codes})
      AND revenue IS NOT NULL
    ORDER BY code, end_date
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()

    if not rows:
        return pd.DataFrame(columns=["code", "rev_consistency", "rev_acceleration",
                                     "rev_quarters"])

    df = pd.DataFrame(rows, columns=["code", "end_date", "revenue"])
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce")
    df = df.dropna(subset=["revenue"])

    results = []
    for code, grp in df.groupby("code"):
        grp = grp.sort_values("end_date").drop_duplicates("end_date", keep="last")
        n = len(grp)

        # YoY: (연,분기) 키 기반으로 정확히 4분기 전과 비교
        # (위치 기반 shift(4)는 분기 결측 시 '4행 전'이 전년 동기가 아니게 됨)
        qidx = pd.PeriodIndex(pd.to_datetime(grp["end_date"]), freq="Q")
        s = pd.Series(grp["revenue"].values, index=qidx)
        s = s[~s.index.duplicated(keep="last")]
        s = s.reindex(pd.period_range(s.index.min(), s.index.max(), freq="Q"))
        yoy = ((s - s.shift(4)) / s.shift(4)).replace([float("inf"), float("-inf")], pd.NA)
        yoy_valid = yoy.dropna().astype(float)

        # 최소 분기 수 미만 or YoY 계산 불가 → NaN 처리
        if n < MIN_REV_QUARTERS or len(yoy_valid) < 2:
            results.append({"code": code, "rev_consistency": None,
                            "rev_acceleration": None, "rev_quarters": n})
            continue

        # 최근 yoy_quarters 개만 사용
        recent = yoy_valid.tail(yoy_quarters)
        consistency = round(float((recent > 0).sum()) / len(recent), 3)

        # 가속도: 최근 4Q 평균 YoY - 전체 평균 YoY (±50% cap으로 극단값 방지)
        all_avg   = yoy_valid.mean()
        recent4   = yoy_valid.tail(4).mean()
        accel_raw = (recent4 - all_avg) * 100  # % 단위로 환산
        accel     = round(max(-REV_ACCEL_CAP, min(REV_ACCEL_CAP, accel_raw)), 2)

        results.append({
            "code":            code,
            "rev_consistency": consistency,
            "rev_acceleration": accel,
            "rev_quarters":    n,
        })

    return pd.DataFrame(results)


def get_operating_margin_ttm(conn, universe_codes, table="quarterly_financials_kr", n_quarters=4):
    """
    quarterly_financials_kr 의 최근 n_quarters 매출/영업이익 합계로 TTM 영업이익률(%) 계산.
    (KR daily_fundamental 에는 operating_margin 컬럼이 없어 매출/영업이익에서 직접 산출)
    """
    if not universe_codes:
        return pd.DataFrame(columns=["code", "operating_margin"])

    in_codes = ", ".join(f"'{c}'" for c in universe_codes)
    sql = f"""
    SELECT code, TO_CHAR(end_date, 'YYYY-MM-DD') AS end_date, revenue, op_income
    FROM {table}
    WHERE code IN ({in_codes})
      AND revenue IS NOT NULL AND op_income IS NOT NULL
    ORDER BY code, end_date
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()
    if not rows:
        return pd.DataFrame(columns=["code", "operating_margin"])

    df = pd.DataFrame(rows, columns=["code", "end_date", "revenue", "op_income"])
    df["revenue"]   = pd.to_numeric(df["revenue"], errors="coerce")
    df["op_income"] = pd.to_numeric(df["op_income"], errors="coerce")

    results = []
    for code, grp in df.groupby("code"):
        grp = grp.sort_values("end_date").drop_duplicates("end_date", keep="last").tail(n_quarters)
        rev_sum = grp["revenue"].sum()
        if len(grp) < n_quarters or rev_sum == 0:
            results.append({"code": code, "operating_margin": None})
            continue
        results.append({"code": code, "operating_margin": round(grp["op_income"].sum() / rev_sum, 4)})

    return pd.DataFrame(results)


def get_roe_debt_equity_kr(conn, universe_codes, table="quarterly_financials_kr", n_quarters=4):
    """
    quarterly_financials_kr 의 net_income/total_equity/total_liabilities 로
    ROE(TTM 순이익/최신 자본총계, %) 와 부채비율(최신 부채총계/자본총계, %) 산출.
    자본총계/부채총계는 재무상태표 스냅샷이라 최신 분기 값을 그대로 사용.
    """
    if not universe_codes:
        return pd.DataFrame(columns=["code", "roe", "debt_to_equity"])

    in_codes = ", ".join(f"'{c}'" for c in universe_codes)
    sql = f"""
    SELECT code, TO_CHAR(end_date, 'YYYY-MM-DD') AS end_date,
           net_income, total_equity, total_liabilities
    FROM {table}
    WHERE code IN ({in_codes})
    ORDER BY code, end_date
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cur.close()
    if not rows:
        return pd.DataFrame(columns=["code", "roe", "debt_to_equity"])

    df = pd.DataFrame(rows, columns=["code", "end_date", "net_income", "total_equity", "total_liabilities"])
    for col in ["net_income", "total_equity", "total_liabilities"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    results = []
    for code, grp in df.groupby("code"):
        grp = grp.sort_values("end_date").drop_duplicates("end_date", keep="last")

        latest_eq = grp["total_equity"].dropna()
        latest_li = grp["total_liabilities"].dropna()
        eq = latest_eq.iloc[-1] if len(latest_eq) else None
        li = latest_li.iloc[-1] if len(latest_li) else None

        ni_recent = grp["net_income"].dropna().tail(n_quarters)
        roe = None
        if eq and eq > 0 and len(ni_recent) == n_quarters:
            roe = round(ni_recent.sum() / eq, 4)

        de = round(li / eq * 100, 2) if (eq and eq > 0 and li is not None) else None

        results.append({"code": code, "roe": roe, "debt_to_equity": de})

    return pd.DataFrame(results)


# ── STEP 3: 가중치 점수 계산 ─────────────────────────────────────────────────

def calc_weighted_factor_score(df, value_col, w_start=W_START, w_step=W_STEP,
                               accel_quarters: int = 4):
    """
    분기별 QoQ 성장률에 선형 가중치 적용 → 종목별 점수 + 일관성 + 가속도 계산

    score       = Σ(QoQ% × weight) / Σ(weight)
    consistency = QoQ > 0 인 분기 수 / 전체 분기 수   (0~1, 높을수록 꾸준)
    acceleration = 최근 N분기 평균 QoQ − 전체 평균 QoQ  (양수=가속, 음수=감속)
    """
    df = df.copy().sort_values(["code", "date_str"])
    df["prev_val"] = df.groupby("code")[value_col].shift(1)
    df["qoq_pct"]  = (df[value_col] - df["prev_val"]) / df["prev_val"] * 100
    df = df.dropna(subset=["qoq_pct", value_col, "prev_val"])

    quarters_sorted = sorted(df["date_str"].unique())
    q_idx_map = {q: i for i, q in enumerate(quarters_sorted)}
    df["q_idx"]        = df["date_str"].map(q_idx_map)
    df["weight"]       = w_start + df["q_idx"] * w_step
    df["weighted_qoq"] = df["qoq_pct"] * df["weight"]

    def _agg(x):
        n   = len(x)
        avg = x["qoq_pct"].mean()
        # 최근 accel_quarters 분기 = q_idx 가장 높은 것들
        recent_n  = min(accel_quarters, n)
        recent_avg = x.nlargest(recent_n, "q_idx")["qoq_pct"].mean()
        return pd.Series({
            "n_quarters":  n,
            "score":       x["weighted_qoq"].sum() / x["weight"].sum(),
            "consistency": round((x["qoq_pct"] > 0).sum() / n, 3),
            "acceleration": round(recent_avg - avg, 2),
        })

    result = (
        df.groupby("code")
        .apply(_agg, include_groups=False)
        .reset_index()
    )

    # 분기별 상세 데이터
    detail = df[["code", "date_str", "yr", "q_num", value_col, "prev_val",
                 "qoq_pct", "weight", "weighted_qoq"]].copy()
    detail["qoq_pct"]      = detail["qoq_pct"].round(2)
    detail["weight"]       = detail["weight"].round(2)
    detail["weighted_qoq"] = detail["weighted_qoq"].round(2)

    return result, detail


# ── 종합 점수 계산 ────────────────────────────────────────────────────────────

def calc_composite_score(result, weights=COMPOSITE_WEIGHTS,
                         inverted=INVERTED_FACTORS):
    """
    각 팩터를 순위(퍼센타일) 정규화(0~1) 후 가중합 → 종합점수 (0~100)
    inverted 팩터는 역방향 정규화 (낮을수록 좋음 → 높은 점수)
    NaN 안전 처리: NaN 팩터는 해당 종목의 가중치 합산에서 제외
    """
    df = result.copy()
    normalized = {}
    for col, w in weights.items():
        if col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce")
        if col == "peg":
            # PEG: 양수만 유효, 상한 cap 적용
            s = s.where(s > 0).clip(upper=PEG_CAP)
        elif col == "debt_to_equity":
            # D/E: 음수(자본잠식/대규모 자사주) → 최악 처리, 상한 cap
            s = s.where(s >= 0, other=DE_CAP).clip(upper=DE_CAP)
        # 순위(퍼센타일) 정규화 — 이상치에 둔감하고 각 팩터를 0~1로 고르게 분포시킴
        # (기존 min-max는 이상치 1개가 스케일을 독식해 나머지를 0 근처로 뭉개고,
        #  그 결과 PER/PBR 같은 팩터가 종합점수에 거의 반영되지 않았음)
        if s.notna().sum() <= 1:
            normalized[col] = pd.Series(0.5, index=df.index)
        else:
            asc = col not in inverted          # 일반=높을수록 1, inverted=낮을수록 1
            normalized[col] = s.rank(ascending=asc, pct=True)

    # NaN-aware 가중 평균
    score      = pd.Series(0.0, index=df.index)
    weight_sum = pd.Series(0.0, index=df.index)
    for col, w in weights.items():
        if col not in normalized:
            continue
        valid       = normalized[col].notna().astype(float)
        score      += normalized[col].fillna(0.0) * w * valid
        weight_sum += w * valid

    df["종합점수"] = (
        score / weight_sum.replace(0, float("nan")) * 100
    ).round(2)
    return df


def apply_gates(result):
    """
    검증불가 팩터(ROE·부채비율)를 '점수'가 아니라 '게이트(통과/탈락 필터)'로 적용.
    값이 분명히 나쁜 종목만 후보에서 제외하고, 값이 없으면(결측) 통과시킨다
    (데이터 없음을 탈락 사유로 삼지 않음). USE_GATES=False 면 원본 그대로 반환.
    반환: (필터된 result, 제외된 종목코드 리스트)
    """
    if not USE_GATES:
        return result, []
    df = result
    keep = pd.Series(True, index=df.index)
    if "roe" in df.columns:
        roe = pd.to_numeric(df["roe"], errors="coerce")
        keep &= ~(roe < GATE_ROE_MIN).fillna(False)
    if "debt_to_equity" in df.columns:
        de = pd.to_numeric(df["debt_to_equity"], errors="coerce")
        keep &= ~(de > GATE_DE_MAX).fillna(False)
    excluded = df.loc[~keep, "code"].tolist()
    return df[keep].reset_index(drop=True), excluded


def compute_safety_columns(result):
    """
    안전(참고지표) 컬럼 추가 — IC 검증상 '안전↔저수익'이라 점수에 합산하지 않는다.
    일별 close 직전 SAFETY_WIN 거래일로 실현변동성·하방편차·최대낙폭(MDD)을 계산하고,
    변동성·하방편차 역순위 평균을 0~100 '안전도'로, 종합점수×안전도 중앙값 기준 '사분면' 라벨 부여.
    실패(데이터 부족 등) 시 원본을 그대로 반환.
    """
    codes = [str(c) for c in result.get("code", pd.Series(dtype=str)).tolist()]
    if not codes:
        return result
    try:
        conn = get_conn()
        cur = conn.cursor()
        inc = ", ".join("'%s'" % c for c in codes)
        cur.execute(
            "SELECT TO_CHAR(date_,'YYYY-MM-DD'), code, close FROM daily_marcap_us "
            "WHERE code IN (%s) AND date_ >= (SELECT MAX(date_) - 400 FROM daily_marcap_us)" % inc)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"  ⚠️  안전도 계산 건너뜀(일별 close 조회 실패): {e!r}")
        return result
    if not rows:
        print("  ⚠️  안전도 계산 건너뜀(일별 close 없음)")
        return result

    d = pd.DataFrame(rows, columns=["d", "code", "close"])
    d["close"] = pd.to_numeric(d["close"], errors="coerce")
    pv = d.pivot_table(index="d", columns="code", values="close").sort_index().tail(SAFETY_WIN)
    r = np.log(pv / pv.shift(1))
    vol, dd, mdd = {}, {}, {}
    for c in pv.columns:
        rc = r[c]
        if rc.notna().sum() < SAFETY_MINVALID:
            continue
        vol[c] = rc.std() * np.sqrt(252)
        dd[c]  = np.sqrt((rc.clip(upper=0) ** 2).mean()) * np.sqrt(252)
        wc = pv[c].dropna()
        mdd[c] = abs(float((wc / wc.cummax() - 1).min()))

    s = pd.DataFrame({"변동성": pd.Series(vol),
                      "하방편차": pd.Series(dd),
                      "MDD": pd.Series(mdd)})
    result = result.merge(s, left_on="code", right_index=True, how="left")

    vr = result["변동성"].rank(pct=True)
    dr = result["하방편차"].rank(pct=True)
    result["안전도"] = (((1 - vr) + (1 - dr)) / 2 * 100).round(1)
    result["변동성"] = (result["변동성"] * 100).round(1)   # 연율 %
    result["MDD"]   = (result["MDD"] * 100).round(1)        # %

    # 사분면: 종합점수(뜨거움) × 안전도 — 중앙값 기준 2×2
    sm = result["종합점수"].median()
    am = result["안전도"].median()

    def _quad(row):
        if pd.isna(row.get("안전도")):
            return ""
        hot  = row["종합점수"] >= sm
        safe = row["안전도"] >= am
        if hot and safe:
            return "강세·안정"
        if hot:
            return "강세·고위험"
        if safe:
            return "약세·안정"
        return "약세·고위험"

    result["사분면"] = result.apply(_quad, axis=1)
    n_ok = int(result["안전도"].notna().sum())
    print(f"  안전도 계산: {n_ok}/{len(result)}종목 "
          f"(변동성·하방편차 역순위 평균 0~100, 점수 미반영·참고지표)")
    return result


# ── Excel 저장 ────────────────────────────────────────────────────────────────

COLOR_HEADER   = "1F4E79"
COLOR_SUB_HDR  = "2E75B6"
COLOR_POS_HIGH = "C6EFCE"   # 연초록: QoQ >= 20%
COLOR_POS_MID  = "FFEB9C"   # 연노랑: 0 <= QoQ < 20%
COLOR_NEG      = "FFC7CE"   # 연빨강: QoQ < 0%
COLOR_ALT_ROW  = "F2F7FF"

# 티어별 색상
TIER_COLORS = {
    "Tier1 (1-10)":    "FFF2CC",   # 노랑
    "Tier2 (11-30)":   "E2EFDA",   # 연초록
    "Tier3 (31-50)":   "DEEAF1",   # 연파랑
    "Tier4 (51-100)":  "F2F2F2",   # 회색
}

thin   = Side(style="thin",   color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def hdr_style(cell, bg=COLOR_HEADER, font_color="FFFFFF", bold=True, size=11):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, color=font_color, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def data_style(cell, alt=False, number_format=None, bg=None):
    fill_color = bg if bg else (COLOR_ALT_ROW if alt else "FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = BORDER
    if number_format:
        cell.number_format = number_format


def color_pct(cell, val):
    try:
        v = float(val)
        if v >= 20:
            cell.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
        elif v >= 0:
            cell.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
        else:
            cell.fill = PatternFill("solid", fgColor=COLOR_NEG)
        cell.font = Font(bold=(abs(v) >= 20))
    except Exception:
        pass


def write_summary_sheet(ws, result, market, years, today_str, name_map=None):
    ws.title = "종합 순위"

    # 제목 행
    ws.merge_cells("A1:Y1")
    title = ws["A1"]
    title.value = (f"Factor Analysis — {market.upper()}  |  "
                   f"최근 {years}년 ({years*4}분기)  |  기준일: {today_str}")
    hdr_style(title, size=13)
    ws.row_dimensions[1].height = 28

    # 가중치 안내
    ws.merge_cells("A2:Y2")
    info = ws["A2"]
    w_max = W_START + (years * 4 - 1) * W_STEP
    _wsum = sum(COMPOSITE_WEIGHTS.values()) or 1.0
    def _wp(k):   # 정규화된 실효 비중(%)
        return COMPOSITE_WEIGHTS.get(k, 0) / _wsum * 100
    info.value = (f"유니버스: 현재 시총 Top{len(result)}  |  "
                  f"선형 가중치: {W_START:.2f}(4년전 Q1) → {w_max:.2f}(최근 Q4)  |  "
                  f"종합점수(IC검증 기반, 실효비중): "
                  f"매출가속도 {_wp('rev_acceleration'):.0f}%  "
                  f"매출성장TTM {_wp('revenue_growth'):.0f}%  "
                  f"매출일관성 {_wp('rev_consistency'):.0f}%  "
                  f"영업이익률 {_wp('operating_margin'):.0f}%  "
                  f"시총성장률 {_wp('시총_성장률'):.0f}%  "
                  f"시총일관성 {_wp('시총_일관성'):.0f}%  "
                  f"거래대금가속도 {_wp('거래대금_가속도'):.0f}%  "
                  f"| PEG·ROE·부채비율 = 게이트(점수 제외)  "
                  f"| 시총가속도·거래대금증가율·거래대금일관성 = 제외(IC≈0)")
    info.fill      = PatternFill("solid", fgColor="EBF3FB")
    info.font      = Font(italic=True, size=10, color="1F4E79")
    info.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # 헤더 (25컬럼) — 마지막 4열은 '안전 참고지표'(점수 미반영)
    headers = [
        "순위", "종목명", "티어",
        "종합점수\n(0~100)",
        "매출 YoY 일관성\n(SEC, 최근8Q)", "매출 YoY 가속도\n(최근4Q-전체평균%)",
        "시총 성장률\n(가중평균 QoQ%)", "시총 가속도\n(최근4Q-전체평균)", "시총 일관성\n(플러스분기비율)",
        "거래대금 증가율\n(가중평균 QoQ%)", "거래대금 가속도\n(최근4Q-전체평균)", "거래대금 일관성\n(플러스분기비율)",
        "시총\n분기수", "매출\n분기수",
        "PER", "PBR", "PEG",
        "ROE\n(%)", "매출 성장률\n(TTM YoY%)", "영업이익률\n(%)", "부채비율\n(D/E)",
        "안전도\n(0~100, 高=안전)", "변동성\n(연율%)", "최대낙폭\n(252d%)", "사분면\n(뜨거움×안전)",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        hdr_style(cell, bg=COLOR_SUB_HDR)
    ws.row_dimensions[3].height = 42

    def _v(row, col):
        v = row.get(col)
        return round(float(v), 2) if pd.notna(v) else None

    # 데이터
    for i, row in result.iterrows():
        r   = i + 4
        alt = (i % 2 == 1)
        tier_color = TIER_COLORS.get(row.get("tier", ""), None)

        # 순위
        rank_cell = ws.cell(row=r, column=1, value=i + 1)
        data_style(rank_cell, alt)
        if i == 0:   rank_cell.font = Font(bold=True, color="FF0000")
        elif i == 1: rank_cell.font = Font(bold=True, color="FF6600")
        elif i == 2: rank_cell.font = Font(bold=True, color="996600")

        # 종목코드 (+ 종목명 병기)
        _name = (name_map or {}).get(row["code"])
        _disp = _name if _name else row["code"]
        code_cell = ws.cell(row=r, column=2, value=_disp)
        data_style(code_cell, alt)
        code_cell.font = Font(bold=True)

        # 티어
        tier_cell = ws.cell(row=r, column=3, value=row.get("tier", ""))
        data_style(tier_cell, alt, bg=tier_color)

        # 종합점수
        v = _v(row, "종합점수")
        c4 = ws.cell(row=r, column=4, value=v)
        data_style(c4, alt, '0.00')
        color_pct(c4, v)   # 점수가 높을수록 초록

        def _consistency_color(cell, v):
            if v is not None:
                if v >= 0.75:   cell.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
                elif v >= 0.5:  cell.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
                else:           cell.fill = PatternFill("solid", fgColor=COLOR_NEG)

        # 매출 YoY 일관성
        v = _v(row, "rev_consistency")
        c5 = ws.cell(row=r, column=5, value=v)
        data_style(c5, alt, '0.00')
        _consistency_color(c5, v)

        # 매출 YoY 가속도
        v = _v(row, "rev_acceleration")
        c6 = ws.cell(row=r, column=6, value=v)
        data_style(c6, alt, '#,##0.00"%"'); color_pct(c6, v)

        # 시총 성장률
        v = _v(row, "시총_성장률")
        c7 = ws.cell(row=r, column=7, value=v)
        data_style(c7, alt, '#,##0.00"%"'); color_pct(c7, v)

        # 시총 가속도
        v = _v(row, "시총_가속도")
        c8 = ws.cell(row=r, column=8, value=v)
        data_style(c8, alt, '#,##0.00"%"'); color_pct(c8, v)

        # 시총 일관성
        v = _v(row, "시총_일관성")
        c9 = ws.cell(row=r, column=9, value=v)
        data_style(c9, alt, '0.00')
        _consistency_color(c9, v)

        # 거래대금 증가율
        v = _v(row, "거래대금_증가율")
        c10 = ws.cell(row=r, column=10, value=v)
        data_style(c10, alt, '#,##0.00"%"'); color_pct(c10, v)

        # 거래대금 가속도
        v = _v(row, "거래대금_가속도")
        c11 = ws.cell(row=r, column=11, value=v)
        data_style(c11, alt, '#,##0.00"%"'); color_pct(c11, v)

        # 거래대금 일관성
        v = _v(row, "거래대금_일관성")
        c12 = ws.cell(row=r, column=12, value=v)
        data_style(c12, alt, '0.00')
        _consistency_color(c12, v)

        # 시총 분기수
        nq_cell = ws.cell(row=r, column=13, value=int(row["n_quarters"]) if pd.notna(row["n_quarters"]) else None)
        data_style(nq_cell, alt)

        # 매출 분기수
        rq = row.get("rev_quarters")
        rq_cell = ws.cell(row=r, column=14, value=int(rq) if pd.notna(rq) and rq is not None else None)
        data_style(rq_cell, alt)

        # PER
        v = _v(row, "per")
        c12 = ws.cell(row=r, column=15, value=v)
        data_style(c12, alt, '0.0')
        if v is not None:
            if v > 50:    c12.fill = PatternFill("solid", fgColor=COLOR_NEG)      # 고평가 경고
            elif v < 15:  c12.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH) # 저평가
            else:         c12.fill = PatternFill("solid", fgColor=COLOR_POS_MID)  # 보통

        # PBR
        v = _v(row, "pbr")
        c16 = ws.cell(row=r, column=16, value=v)
        data_style(c16, alt, '0.00')
        if v is not None:
            if v > 10:   c16.fill = PatternFill("solid", fgColor=COLOR_NEG)
            elif v < 2:  c16.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            else:        c16.fill = PatternFill("solid", fgColor=COLOR_POS_MID)

        # PEG (낮을수록 저평가: <1 초록, 1~15 노랑, >15 빨강, None 회색)
        v = _v(row, "peg")
        c17 = ws.cell(row=r, column=17, value=v)
        data_style(c17, alt, '0.00')
        if v is not None:
            if v < 1:    c17.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            elif v < 15: c17.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
            else:        c17.fill = PatternFill("solid", fgColor=COLOR_NEG)
        else:
            c17.fill = PatternFill("solid", fgColor="E0E0E0")

        # ROE % (소수→퍼센트 변환, 높을수록 좋음: >15% 초록, 0~15% 노랑, 음수 빨강)
        roe_raw = row.get("roe")
        roe_pct = round(float(roe_raw) * 100, 1) if pd.notna(roe_raw) else None
        c18 = ws.cell(row=r, column=18, value=roe_pct)
        data_style(c18, alt, '0.0"%"')
        if roe_pct is not None:
            if roe_pct >= 15:  c18.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            elif roe_pct >= 0: c18.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
            else:              c18.fill = PatternFill("solid", fgColor=COLOR_NEG)
        else:
            c18.fill = PatternFill("solid", fgColor="E0E0E0")

        # 매출 성장률 % (TTM YoY, 소수→퍼센트, 높을수록 좋음)
        rg_raw = row.get("revenue_growth")
        rg_pct = round(float(rg_raw) * 100, 1) if pd.notna(rg_raw) else None
        c19 = ws.cell(row=r, column=19, value=rg_pct)
        data_style(c19, alt, '0.0"%"')
        color_pct(c19, rg_pct)
        if rg_pct is None:
            c19.fill = PatternFill("solid", fgColor="E0E0E0")

        # 영업이익률 % (소수→퍼센트, 높을수록 좋음: >20% 초록, 0~20% 노랑, 음수 빨강)
        om_raw = row.get("operating_margin")
        om_pct = round(float(om_raw) * 100, 1) if pd.notna(om_raw) else None
        c_om = ws.cell(row=r, column=20, value=om_pct)
        data_style(c_om, alt, '0.0"%"')
        if om_pct is not None:
            if om_pct >= 20:  c_om.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            elif om_pct >= 0: c_om.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
            else:             c_om.fill = PatternFill("solid", fgColor=COLOR_NEG)
        else:
            c_om.fill = PatternFill("solid", fgColor="E0E0E0")

        # 부채비율 D/E (낮을수록 좋음: <100 초록, 100~300 노랑, >300 빨강, 음수=자본잠식 빨강)
        de_raw = row.get("debt_to_equity")
        de_val = round(float(de_raw), 1) if pd.notna(de_raw) else None
        c21 = ws.cell(row=r, column=21, value=de_val)
        data_style(c21, alt, '0.0')
        if de_val is not None:
            if de_val < 0:     c21.fill = PatternFill("solid", fgColor=COLOR_NEG)
            elif de_val < 100: c21.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            elif de_val < 300: c21.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
            else:              c21.fill = PatternFill("solid", fgColor=COLOR_NEG)
        else:
            c21.fill = PatternFill("solid", fgColor="E0E0E0")

        # ── 안전(참고지표; 점수 미반영) ───────────────────────────────
        # 안전도 (0~100, 높을수록 저변동=안전)
        sv = _v(row, "안전도")
        c22 = ws.cell(row=r, column=22, value=sv)
        data_style(c22, alt, '0.0'); color_pct(c22, sv)
        if sv is None:
            c22.fill = PatternFill("solid", fgColor="E0E0E0")

        # 변동성 (연율%, 낮을수록 안전)
        vv = _v(row, "변동성")
        c23 = ws.cell(row=r, column=23, value=vv)
        data_style(c23, alt, '0.0"%"')
        if vv is not None:
            if vv < 30:   c23.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            elif vv < 60: c23.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
            else:         c23.fill = PatternFill("solid", fgColor=COLOR_NEG)
        else:
            c23.fill = PatternFill("solid", fgColor="E0E0E0")

        # 최대낙폭 MDD (252d %, 낮을수록 안전)
        mv = _v(row, "MDD")
        c24 = ws.cell(row=r, column=24, value=mv)
        data_style(c24, alt, '0.0"%"')
        if mv is not None:
            if mv < 20:   c24.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)
            elif mv < 40: c24.fill = PatternFill("solid", fgColor=COLOR_POS_MID)
            else:         c24.fill = PatternFill("solid", fgColor=COLOR_NEG)
        else:
            c24.fill = PatternFill("solid", fgColor="E0E0E0")

        # 사분면 (뜨거움×안전)
        qd = row.get("사분면", "") or ""
        c25 = ws.cell(row=r, column=25, value=qd)
        data_style(c25, alt)
        qcolor = {"강세·안정": COLOR_POS_HIGH, "강세·고위험": "FFD9B3",
                  "약세·안정": COLOR_POS_MID, "약세·고위험": COLOR_NEG}.get(qd)
        if qcolor:
            c25.fill = PatternFill("solid", fgColor=qcolor)

    # 컬럼 너비 (A~Y, 25열; V~Y = 안전 참고지표)
    for col, w in zip("ABCDEFGHIJKLMNOPQRSTUVWXY",
                      [8, 24, 18, 12, 20, 20, 22, 20, 18, 22, 20, 18, 9, 9, 12, 12, 12, 12, 14, 13, 12,
                       13, 11, 12, 16]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A4"


def write_detail_sheet(ws, detail, value_col, sheet_name, label, score_col):
    ws.title = sheet_name

    col_names = ["종목코드", "분기말일", "연도", "분기",
                 label, f"전분기 {label}", "QoQ(%)", "가중치", "가중 QoQ"]
    for c, h in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=c, value=h)
        hdr_style(cell, bg=COLOR_SUB_HDR)
    ws.row_dimensions[1].height = 20

    # 종목별 최종 점수 기준으로 정렬: 점수 높은 종목 먼저, 같은 종목 내에서는 날짜 순
    detail = (
        detail
        .sort_values([score_col, "date_str"], ascending=[False, True])
        .reset_index(drop=True)
    )

    for i, row in detail.iterrows():
        r   = i + 2
        alt = (i % 2 == 1)
        vals = [
            row["code"],
            row["date_str"],
            int(row["yr"]),
            int(row["q_num"]),
            round(float(row[value_col]) / 1e9, 2) if pd.notna(row[value_col]) else None,
            round(float(row["prev_val"]) / 1e9, 2) if pd.notna(row["prev_val"]) else None,
            round(float(row["qoq_pct"]),     2)    if pd.notna(row["qoq_pct"])  else None,
            round(float(row["weight"]),      2),
            round(float(row["weighted_qoq"]),2)    if pd.notna(row["weighted_qoq"]) else None,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            data_style(cell, alt)
            if c == 7:   # QoQ%
                color_pct(cell, v)

    for col, w in zip("ABCDEFGHI", [12, 14, 8, 8, 14, 14, 12, 10, 12]):
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"


def write_guide_sheet(ws):
    """팩터 설명 시트"""
    ws.title = "팩터 설명"

    # 제목
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "Factor Analysis — 지표 설명"
    t.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    t.font      = Font(bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 헤더
    for c, h in enumerate(["지표", "계산식", "의미", "해석 기준"], 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill      = PatternFill("solid", fgColor=COLOR_SUB_HDR)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[2].height = 20

    rows = [
        # ── 매출 펀더멘탈 ────────────────────────────────────────────────────────
        ("매출 YoY 일관성",
         "최근 8분기 중 전년 동기 대비\n매출이 증가한 분기 수 ÷ 8  (0~1)\n출처: SEC EDGAR 분기 재무제표\n※ 분기 데이터 12분기 미만이면 NaN",
         "시장 상황과 무관하게 매출이 꾸준히 성장하는지 측정.\n"
         "YoY(전년 동기 비교)를 쓰기 때문에 계절성 제거됨.\n"
         "싸이클 종목과 구조적 성장주를 구분하는 핵심 지표.\n"
         "ex) NVDA가 AI 테마로 8분기 연속 성장 → 1.0\n"
         "※ 12분기(3년) 미만 신생·분사 종목은 역사 부족으로 제외 (NaN)",
         "높을수록 좋음 (0~1)\n1.0: 8분기 전부 YoY 성장\n0.75 이상: 꾸준한 구조적 성장\n"
         "0.5 미만: 싸이클 또는 사업 변동성 큼\n"
         "NaN: 12분기 미만 데이터 — 가중치 제외\n※ 가중치 15% — 가장 중요한 단일 팩터"),

        ("매출 YoY 가속도",
         "최근 4분기 평균 YoY 성장률\n− 전체 분기 평균 YoY 성장률 (%)\n출처: SEC EDGAR 분기 재무제표\n※ ±50% cap 적용",
         "성장이 점점 빨라지고 있는지(가속) 측정.\n"
         "같은 일관성이라도 최근에 더 빠르게 성장 중이면 우위.\n"
         "ex) 전체 평균 YoY 20%, 최근 4Q 평균 YoY 35% → 가속도 +15%\n"
         "※ 극단값(±50% 초과)은 cap 처리하여 정규화 왜곡 방지",
         "양수: 최근 성장 가속 (좋음)\n음수: 최근 성장 둔화 (주의)\n"
         "0에 가까우면 일정한 성장 속도 유지\n"
         f"±{REV_ACCEL_CAP:.0f}% 범위 내로 제한"),

        # ── 모멘텀 ──────────────────────────────────────────────────────────────
        ("시총 성장률",
         "분기별 시가총액 QoQ% 의 가중평균\n(최근 분기일수록 가중치 높음)",
         "최근 몇 년간 시가총액이 얼마나 꾸준히 성장해왔는지 측정.\n"
         "단순 현재 시총 크기가 아니라 '성장 속도'를 봄.",
         "높을수록 좋음\n20% 이상: 강한 성장\n0~20%: 완만한 성장\n음수: 시총 감소"),

        ("시총 가속도",
         "최근 4분기 평균 QoQ%  −  전체 분기 평균 QoQ%",
         "성장이 점점 빨라지고 있는지(가속) 또는 느려지고 있는지(감속)를 측정.\n"
         "같은 성장률이라도 최근에 더 빠르게 성장 중이면 우위.",
         "양수: 최근 성장 가속 중 (좋음)\n음수: 최근 성장 둔화 중 (주의)"),

        ("시총 일관성",
         "QoQ > 0인 분기 수  ÷  전체 분기 수  (0~1)",
         "성장이 얼마나 꾸준한지 측정. 한두 분기에만 튀지 않고\n"
         "지속적으로 성장해왔는지 확인.",
         "1.0: 모든 분기 성장\n0.75 이상: 꾸준한 성장\n0.5 미만: 등락 심함"),

        ("거래대금 증가율",
         "분기별 거래대금 합계 QoQ% 의 가중평균",
         "투자자들의 관심·매매 활동이 늘고 있는지 측정.\n"
         "시총 성장과 함께 거래대금도 늘면 신뢰도 높음.",
         "높을수록 좋음\n시총 성장률과 함께 보면 더 유효"),

        ("거래대금 가속도",
         "최근 4분기 평균 QoQ%  −  전체 분기 평균 QoQ%",
         "거래 활동이 최근 들어 더 활발해지고 있는지 측정.",
         "양수: 최근 관심 증가\n음수: 관심 식는 중"),

        ("거래대금 일관성",
         "QoQ > 0인 분기 수  ÷  전체 분기 수  (0~1)",
         "거래대금이 꾸준히 증가해왔는지 확인.",
         "시총 일관성과 동일한 기준 적용"),

        # ── 종합점수 ─────────────────────────────────────────────────────────────
        ("종합점수",
         "각 팩터를 순위(퍼센타일)로 정규화 후\nNaN-aware 가중합 × 100",
         "IC 검증(33분기) 기반으로 가중치를 재설계.\n"
         "전·후반 부호가 안정적이었던 매출(SEC) 계열로 가볍게 틸트하고,\n"
         "IC≈0/부호반전 모멘텀은 점수에서 제외, PEG·ROE·부채비율은 게이트로 강등.",
         (lambda ws: f"실효비중:\n"
          f"매출가속도 {COMPOSITE_WEIGHTS.get('rev_acceleration',0)/ws*100:.0f}%\n"
          f"매출성장TTM {COMPOSITE_WEIGHTS.get('revenue_growth',0)/ws*100:.0f}%\n"
          f"매출일관성 {COMPOSITE_WEIGHTS.get('rev_consistency',0)/ws*100:.0f}%\n"
          f"영업이익률 {COMPOSITE_WEIGHTS.get('operating_margin',0)/ws*100:.0f}%\n"
          f"+ 잔여 모멘텀 소량\n"
          f"PEG·ROE·부채비율 = 게이트(점수 제외)"
          )(sum(COMPOSITE_WEIGHTS.values()) or 1.0)),

        # ── 밸류에이션 ────────────────────────────────────────────────────────────
        ("PER",
         "주가  ÷  주당순이익(EPS)\n(Trailing: 최근 12개월 실적 기준)",
         "현재 주가가 이익 대비 얼마나 비싼지 측정.\n"
         "적자 기업은 PER 계산 불가 → 빈칸.",
         "낮을수록 저평가\n15 미만: 저평가 (초록)\n15~50: 적정 (노랑)\n50 초과: 고평가 (빨강)\n"
         "단, 업종별 평균이 다르므로 절대값보다 상대 비교 중요"),

        ("PBR",
         "주가  ÷  주당순자산(BPS)",
         "현재 주가가 회사 자산 대비 얼마나 비싼지 측정.\n"
         "음수는 자본잠식(부채 > 자산) 상태.",
         "낮을수록 저평가\n2 미만: 저평가 (초록)\n2~10: 적정 (노랑)\n10 초과: 고평가 (빨강)\n"
         "자산 경량 기술기업은 PBR이 높아도 정상"),

        ("ROE",
         "순이익 ÷ 자기자본 × 100\nyfinance: returnOnEquity (소수 → ×100 = %)",
         "내 돈(자기자본) 100원으로 얼마를 벌었는지 측정.\n"
         "높을수록 경영 효율이 좋고 주주에게 유리함.\n"
         "워런 버핏이 가장 중요시하는 지표 중 하나.",
         "높을수록 좋음\n15% 이상: 우량 (초록)\n0~15%: 보통 (노랑)\n음수: 적자 (빨강)\n"
         "단, 자사주 매입 많으면 자기자본 감소로 ROE 과대 표시될 수 있음"),

        ("매출 성장률",
         "YoY 매출 증가율\nyfinance: revenueGrowth (소수 → ×100 = %)",
         "1년 전 대비 매출이 얼마나 늘었는지.\n"
         "PEG는 이익 성장이지만, 매출 성장은 이익이 아직 없어도\n"
         "사업이 커지고 있는지 확인할 수 있음.\n"
         "성장 초기 기업 판단에 특히 유용.",
         "높을수록 좋음\n20% 이상: 강한 성장\n0~20%: 완만한 성장\n음수: 매출 감소"),

        ("영업이익률",
         "영업이익 ÷ 매출 × 100\nyfinance: operatingMargins (소수 → ×100 = %)",
         "본업(영업)에서 매출의 몇 %를 이익으로 남기는지.\n"
         "PER은 '비싸냐 싸냐'이고, 영업이익률은 '사업 자체가 좋냐'.\n"
         "높을수록 경쟁 우위 및 가격 결정력이 강함.\n"
         "예: NVDA 60%대, 일반 제조업 5~10%",
         "높을수록 좋음\n20% 이상: 고수익 사업 (초록)\n0~20%: 보통 (노랑)\n음수: 영업 손실 (빨강)"),

        ("부채비율 (D/E)",
         "총부채 ÷ 자기자본\nyfinance: debtToEquity",
         "부채가 자기자본의 몇 배인지 측정하는 재무 건전성 지표.\n"
         "낮을수록 재무 리스크 적음 (역순 정규화 적용).\n"
         "음수: 자기자본 자체가 음수 (대규모 손실 또는 자사주 매입 과다)\n"
         "※ Apple, ABBV 등 자사주 매입이 많은 기업은 의도적으로\n"
         "   자기자본이 낮아 D/E가 높거나 음수로 나올 수 있음.",
         "낮을수록 좋음 (역순 정규화)\n100 미만: 건전 (초록)\n100~300: 주의 (노랑)\n300 초과: 고위험 (빨강)\n음수: 자본잠식 (빨강)\n종합점수 내 가중치 낮음 (4%) — 참고용"),

        ("PEG",
         "PER  ÷  EPS 연간 성장률(%)\n예) PER=30, 성장률=30% → PEG=1.0",
         "성장 속도를 감안했을 때 주가가 적정한지 측정.\n"
         "PER만 보면 성장주가 무조건 고평가처럼 보이지만\n"
         "성장률이 높으면 PEG는 낮게 나옴.\n\n"
         "※ EPS 성장률 양수인 경우만 '낮을수록 저평가' 성립.\n"
         "  이익 감소(음수 성장)는 PEG 최하 점수 부여.\n"
         "  PER 없음(적자)은 데이터 부재로 가중치에서 제외.",
         "낮을수록 저평가 (역순 정규화 적용)\n"
         "1 미만: 성장 대비 저평가 (초록)\n"
         "1~15: 적정\n"
         "15 초과: 성장 대비 고평가 (빨강)\n"
         f"이익 감소: {PEG_CAP:.0f}(최하) 처리\n"
         "빈칸(N/A): 적자 기업 — 가중치 제외"),
    ]

    # 구분선용 섹션 헤더
    section_rows = {
        0:  "📊 매출 펀더멘탈 팩터 (SEC EDGAR 분기 히스토리)",
        2:  "📈 모멘텀 팩터",
        8:  "🏆 종합점수",
        9:  "💰 밸류에이션 팩터 (PER / PBR)",
        11: "📊 펀더멘털 품질 팩터 (ROE / 매출성장률TTM / 영업이익률 / 부채비율)",
        15: "🔗 복합 지표 (PEG)",
    }

    data_row = 3
    for idx, (name, formula, meaning, criteria) in enumerate(rows):
        # 섹션 헤더 삽입
        if idx in section_rows:
            ws.merge_cells(f"A{data_row}:D{data_row}")
            sec = ws.cell(row=data_row, column=1, value=section_rows[idx])
            sec.fill      = PatternFill("solid", fgColor="D6E4F0")
            sec.font      = Font(bold=True, size=10, color="1F4E79")
            sec.alignment = Alignment(horizontal="left", vertical="center",
                                      indent=1)
            sec.border    = BORDER
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        vals = [name, formula, meaning, criteria]
        heights = [max(v.count("\n") + 1 for v in vals) * 14 + 6]
        ws.row_dimensions[data_row].height = max(heights[0], 18)

        alt = (idx % 2 == 1)
        bg  = "F7FBFF" if alt else "FFFFFF"
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=data_row, column=c, value=v)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(size=10, bold=(c == 1))
            cell.alignment = Alignment(horizontal="left" if c > 1 else "center",
                                       vertical="top", wrap_text=True)
            cell.border    = BORDER
        data_row += 1

    # 컬럼 너비
    for col, w in zip("ABCD", [18, 32, 52, 40]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"


TREND_TOP_N = 50   # 시총 추이 탭에 담을 현재 상위 종목 수
TREND_HORIZONS = [("1주", 7), ("1개월", 30), ("3개월", 91),
                  ("6개월", 182), ("1년", 365), ("3년", 1095), ("5년", 1825)]
TREND_LABELS = [h[0] for h in TREND_HORIZONS]


def compute_trend_data(market="us", top_n=TREND_TOP_N):
    """현재 시총 Top N 종목의 과거(1주·1개월·분기·6개월·1년) 대비 종가 수익률·순위변화."""
    marcap_tbl = "daily_marcap" if market == "kr" else "daily_marcap_us"
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(f"SELECT TO_CHAR(MAX(date_),'YYYY-MM-DD') FROM {marcap_tbl}")
        today = cur.fetchone()[0]
        ref = {"현재": today}
        for label, days in TREND_HORIZONS:
            target = (pd.Timestamp(today) - pd.Timedelta(days=days)).strftime("%Y-%m-%d")
            cur.execute(f"SELECT TO_CHAR(MAX(date_),'YYYY-MM-DD') FROM {marcap_tbl} "
                        f"WHERE date_ <= TO_DATE('{target}','YYYY-MM-DD')")
            ref[label] = cur.fetchone()[0]
        # 현재 Top N (GOOG 등 중복 클래스 제외 후 N개) — 버퍼로 넉넉히 조회 후 잘라냄
        cur.execute(f"SELECT code, rank FROM {marcap_tbl} "
                    f"WHERE date_ = TO_DATE('{today}','YYYY-MM-DD') AND rank <= {top_n + 10} "
                    f"ORDER BY rank")
        toprows = [(c, r) for (c, r) in cur.fetchall() if c not in EXCLUDE_CODES][:top_n]
        codes = [c for c, _ in toprows]
        if not codes:
            cur.close(); conn.close(); return None, None
        uniq_dates = sorted(set(ref.values()))
        ins = ", ".join(f"TO_DATE('{d}','YYYY-MM-DD')" for d in uniq_dates)
        inc = ", ".join("'%s'" % c for c in codes)
        cur.execute(f"SELECT TO_CHAR(date_,'YYYY-MM-DD'), code, close, marcap, rank "
                    f"FROM {marcap_tbl} WHERE date_ IN ({ins}) AND code IN ({inc})")
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        print(f"  ⚠️  시총 추이 계산 건너뜀: {e!r}")
        return None, None

    df = pd.DataFrame(rows, columns=["d", "code", "close", "marcap", "rank"])
    for c in ["close", "marcap", "rank"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    close_pv = df.pivot_table(index="code", columns="d", values="close")
    mc_pv    = df.pivot_table(index="code", columns="d", values="marcap")
    rank_pv  = df.pivot_table(index="code", columns="d", values="rank")
    t = ref["현재"]

    def _get(pv, code, d):
        if code in pv.index and d in pv.columns:
            v = pv.at[code, d]
            return v if pd.notna(v) else None
        return None

    recs = []
    for code, rk in toprows:
        cur_close = _get(close_pv, code, t)
        mc = _get(mc_pv, code, t)
        rec = {"순위": int(rk), "종목": code,
               "시총($B)": round(mc / 1e9, 1) if mc else None}
        for label in TREND_LABELS:
            past = _get(close_pv, code, ref[label])
            rec[label] = round((cur_close / past - 1) * 100, 1) if (cur_close and past) else None
        rq = _get(rank_pv, code, ref["3개월"])
        rec["분기전순위"] = int(rq) if rq is not None else None
        rec["순위Δ"] = (int(rq) - int(rk)) if rq is not None else None   # +면 순위 상승
        recs.append(rec)
    return pd.DataFrame(recs), ref


def write_trend_sheet(ws, trend_df, ref, market):
    """현재 시총 Top N 종목의 기간별 종가 수익률 표 + 막대그래프."""
    from openpyxl.chart import BarChart, Reference
    ws.title = "시총Top추이"
    n = len(trend_df)

    ws.merge_cells("A1:L1")
    ttl = ws["A1"]
    ttl.value = (f"미국 시총 Top{n} 추이 (NASDAQ·NYSE 대형주, 종가 기준 수익률)"
                 f"   |   현재 기준일: {ref['현재']}")
    hdr_style(ttl, size=13)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:L2")
    info = ws["A2"]
    info.value = ("비교 기준일 — " + "  ".join(f"{k}: {ref[k]}" for k in TREND_LABELS)
                  + "   |   값 = (현재 종가 ÷ 과거 종가 − 1), 분할보정 종가")
    info.fill = PatternFill("solid", fgColor="EBF3FB")
    info.font = Font(italic=True, size=10, color="1F4E79")
    info.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["현재\n순위", "종목", "시총\n($B)",
               "1주\n(%)", "1개월\n(%)", "3개월\n(%)", "6개월\n(%)",
               "1년\n(%)", "3년\n(%)", "5년\n(%)", "3개월전\n순위", "순위\n변화"]
    for c, h in enumerate(headers, 1):
        hdr_style(ws.cell(3, c, h), bg=COLOR_SUB_HDR)
    ws.row_dimensions[3].height = 34

    # 기간 컬럼: 1주(4) … 5년(10)
    pct_cols = {label: 4 + i for i, label in enumerate(TREND_LABELS)}
    rank_col, chg_col = 11, 12
    for i, (_, row) in enumerate(trend_df.iterrows()):
        r = i + 4
        alt = (i % 2 == 1)
        data_style(ws.cell(r, 1, int(row["순위"])), alt)
        cc = ws.cell(r, 2, row["종목"]); data_style(cc, alt); cc.font = Font(bold=True)
        mc = row["시총($B)"]
        data_style(ws.cell(r, 3, mc if pd.notna(mc) else None), alt, '#,##0.0')
        for label, col in pct_cols.items():
            v = row[label]
            cell = ws.cell(r, col, v if pd.notna(v) else None)
            data_style(cell, alt, '0.0"%"')
            if pd.notna(v):
                color_pct(cell, v)
            else:
                cell.fill = PatternFill("solid", fgColor="E0E0E0")   # 상장 전 등 데이터 없음
        rq = row["분기전순위"]
        data_style(ws.cell(r, rank_col, int(rq) if pd.notna(rq) else None), alt)
        dch = row["순위Δ"]
        dcell = ws.cell(r, chg_col, int(dch) if pd.notna(dch) else None)
        data_style(dcell, alt)
        if pd.notna(dch):
            if dch > 0:   dcell.fill = PatternFill("solid", fgColor=COLOR_POS_HIGH)   # 순위 상승
            elif dch < 0: dcell.fill = PatternFill("solid", fgColor=COLOR_NEG)         # 순위 하락

    last = n + 3
    for col, w in zip("ABCDEFGHIJKL", [7, 10, 11, 8, 8, 8, 8, 8, 8, 8, 11, 9]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # 막대그래프 3개: 1개월 · 3개월 · 1년 (단기·중기·장기 추이)
    cats = Reference(ws, min_col=2, min_row=4, max_row=last)
    chart_specs = [(f"1개월 수익률 (%) — 현재 Top{n}", pct_cols["1개월"], f"A{last + 2}"),
                   (f"3개월 수익률 (%) — 현재 Top{n}", pct_cols["3개월"], f"A{last + 22}"),
                   (f"1년 수익률 (%) — 현재 Top{n}",   pct_cols["1년"],  f"A{last + 42}")]
    for title, col, anchor in chart_specs:
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        data = Reference(ws, min_col=col, min_row=3, max_row=last)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.height = 9
        ch.width = 38
        ch.legend = None
        ch.gapWidth = 40
        ws.add_chart(ch, anchor)


def save_excel(result, f1_detail, f2_detail, market, years, out_path,
               trend_df=None, ref_dates=None, name_map=None):
    wb = Workbook()
    today_str = date.today().strftime("%Y-%m-%d")

    write_summary_sheet(wb.active, result, market, years, today_str, name_map)
    if trend_df is not None and len(trend_df):
        write_trend_sheet(wb.create_sheet("시총Top추이", 1), trend_df, ref_dates, market)
    write_detail_sheet(wb.create_sheet(), f1_detail, "marcap",
                       "시총_상세", "시총($B)", "시총_성장률")
    write_detail_sheet(wb.create_sheet(), f2_detail, "total_amount",
                       "거래대금_상세", "거래대금($B)", "거래대금_증가율")
    write_guide_sheet(wb.create_sheet())

    wb.save(out_path)
    print(f"  Excel 저장: {out_path}")


# ── 메인 ──────────────────────────────────────────────────────────────────────

def run_factor_analysis(market="us", years=YEARS, top_n=TOP_N):
    marcap_tbl = "daily_marcap"  if market == "kr" else "daily_marcap_us"
    price_tbl  = "daily_price"   if market == "kr" else "daily_price_us"
    n_quarters = years * 4
    w_max      = W_START + (n_quarters - 1) * W_STEP

    print("=" * 60)
    print(f"  Factor Analysis -- {market.upper()}")
    print(f"  기간   : 최근 {years}년 ({n_quarters}분기)")
    print(f"  유니버스: 현재 시총 상위 {top_n}개 (고정)")
    print(f"  가중치 : {W_START:.2f}(oldest) ~ {w_max:.2f}(newest), step={W_STEP}")
    print("=" * 60)

    conn = get_conn()
    print("  DB 연결 성공\n")
    name_map = get_name_map(conn, "ticker_master_us" if market == "us" else "ticker_master")

    # Step 0: 현재 유니버스 확정
    print("[0] 현재 유니버스 조회 (고정 Top N)...")
    universe_df = get_current_universe(conn, top_n=top_n, table=marcap_tbl)
    if EXCLUDE_CODES:
        removed = universe_df[universe_df["code"].isin(EXCLUDE_CODES)]["code"].tolist()
        universe_df = universe_df[~universe_df["code"].isin(EXCLUDE_CODES)].reset_index(drop=True)
        if removed:
            print(f"  중복 클래스 제거: {removed}")
    # 제외 후에도 정확히 top_n 유지 (버퍼 +10 조회분에서 보충)
    universe_df = universe_df.head(top_n).reset_index(drop=True)
    universe_codes = universe_df["code"].tolist()
    print(f"  {len(universe_codes)}개 종목 확정")
    print(f"  상위 10: {', '.join(universe_codes[:10])}")

    # Step 1: 분기말 날짜
    print("\n[1] 분기말 날짜 조회...")
    qdates = get_quarter_end_dates(conn, years=years, table=marcap_tbl)
    print(f"  {len(qdates)}개 분기:")
    for qd, yr, q in qdates:
        print(f"     {yr} Q{q}  ({str(qd)[:10]})")

    # Step 2-A: 시총 성장률
    print(f"\n[2] 시총 성장률 분기말 스냅샷...")
    df_marcap = get_quarterly_marcap(conn, qdates, universe_codes, table=marcap_tbl)
    print(f"  {len(df_marcap)}행 / {df_marcap['code'].nunique()}종목")
    f1_score, f1_detail = calc_weighted_factor_score(df_marcap, "marcap")
    f1_score = f1_score.rename(columns={
        "score":        "시총_성장률",
        "consistency":  "시총_일관성",
        "acceleration": "시총_가속도",
    })

    # Step 2-B: 거래대금 증가율
    print(f"\n[3] 거래대금 증가율 분기별 합계...")
    df_amount = get_quarterly_amount(conn, qdates, universe_codes, price_table=price_tbl)
    print(f"  {len(df_amount)}행 / {df_amount['code'].nunique()}종목")
    f2_score, f2_detail = calc_weighted_factor_score(df_amount, "total_amount")
    f2_score = f2_score.rename(columns={
        "score":        "거래대금_증가율",
        "consistency":  "거래대금_일관성",
        "acceleration": "거래대금_가속도",
    })

    # Step 2-C: 매출 YoY 일관성 / 가속도
    rev_tbl = "quarterly_financials_us" if market == "us" else "quarterly_financials_kr"
    print(f"\n[3-C] 매출 YoY 일관성/가속도 ({rev_tbl})...")
    rev_factor_df = get_revenue_factors(conn, universe_codes, table=rev_tbl)
    ok = rev_factor_df["rev_consistency"].notna().sum() if not rev_factor_df.empty else 0
    print(f"  {len(rev_factor_df)}종목 / 일관성 데이터 보유: {ok}종목")

    # Step 3: 펀더멘털 (PER/PBR/PEG)
    fund_df = pd.DataFrame()
    if market == "us":
        print(f"\n[4] 펀더멘털 조회 (daily_fundamental_us)...")
        fund_sql = """
        SELECT code, per, pbr, div, eps, eps_growth,
               roe, revenue_growth, debt_to_equity, operating_margin
        FROM daily_fundamental_us
        WHERE date_ = (SELECT MAX(date_) FROM daily_fundamental_us)
          AND code IN ({})
        """.format(", ".join(f"'{c}'" for c in universe_codes))
        cur = conn.cursor()
        cur.execute(fund_sql)
        rows = cur.fetchall()
        cur.close()
        if rows:
            fund_df = pd.DataFrame(rows, columns=[
                "code", "per", "pbr", "div", "eps", "eps_growth",
                "roe", "revenue_growth", "debt_to_equity", "operating_margin"
            ])
            for col in fund_df.columns[1:]:
                fund_df[col] = pd.to_numeric(fund_df[col], errors="coerce")
            print(f"  {len(fund_df)}건 조회됨  "
                  f"(roe: {fund_df['roe'].notna().sum()}  "
                  f"rev_growth: {fund_df['revenue_growth'].notna().sum()}  "
                  f"op_margin: {fund_df['operating_margin'].notna().sum()}  "
                  f"d/e: {fund_df['debt_to_equity'].notna().sum()})")
        else:
            print("  ⚠️  데이터 없음 — etl_fundamental_us.py 먼저 실행 필요")
    elif market == "kr":
        # KR daily_fundamental 에는 eps_growth/revenue_growth 컬럼이 없음 (EPS 추세 산출 미구축)
        # ROE/부채비율은 quarterly_financials_kr 의 net_income/total_equity/total_liabilities 로 산출
        # (게이트 전용 — US와 동일하게 종합점수에는 미반영)
        print(f"\n[4] 펀더멘털 조회 (daily_fundamental PER/PBR/DIV/EPS + DART ROE/부채비율)...")
        fund_sql = """
        SELECT code, per, pbr, div, eps
        FROM daily_fundamental
        WHERE date_ = (SELECT MAX(date_) FROM daily_fundamental)
          AND code IN ({})
        """.format(", ".join(f"'{c}'" for c in universe_codes))
        cur = conn.cursor()
        cur.execute(fund_sql)
        rows = cur.fetchall()
        cur.close()
        if rows:
            fund_df = pd.DataFrame(rows, columns=["code", "per", "pbr", "div", "eps"])
            for col in fund_df.columns[1:]:
                fund_df[col] = pd.to_numeric(fund_df[col], errors="coerce")
            om_df = get_operating_margin_ttm(conn, universe_codes)
            fund_df = fund_df.merge(om_df, on="code", how="left")
            roe_df = get_roe_debt_equity_kr(conn, universe_codes)
            fund_df = fund_df.merge(roe_df, on="code", how="left")
            for col in ["eps_growth", "revenue_growth"]:
                fund_df[col] = None
            print(f"  {len(fund_df)}건 조회됨  "
                  f"(per: {fund_df['per'].notna().sum()}  "
                  f"op_margin: {fund_df['operating_margin'].notna().sum()}  "
                  f"roe: {fund_df['roe'].notna().sum()}  "
                  f"d/e: {fund_df['debt_to_equity'].notna().sum()})")
        else:
            print("  ⚠️  데이터 없음 (daily_fundamental)")

    conn.close()

    # 결합 + 티어 부여
    result = (
        f1_score[["code", "n_quarters", "시총_성장률", "시총_일관성", "시총_가속도"]]
        .merge(f2_score[["code", "거래대금_증가율", "거래대금_일관성", "거래대금_가속도"]], on="code", how="outer")
        .merge(universe_df[["code", "current_rank"]], on="code", how="left")
    )
    # 매출 팩터 병합
    if not rev_factor_df.empty:
        result = result.merge(
            rev_factor_df[["code", "rev_consistency", "rev_acceleration", "rev_quarters"]],
            on="code", how="left"
        )
    else:
        result["rev_consistency"]  = None
        result["rev_acceleration"] = None
        result["rev_quarters"]     = None
    # 펀더멘털 병합
    if not fund_df.empty:
        result = result.merge(
            fund_df[["code", "per", "pbr", "eps", "eps_growth",
                     "roe", "revenue_growth", "debt_to_equity", "operating_margin"]],
            on="code", how="left"
        )
        for col in ["per", "pbr", "eps_growth", "roe", "revenue_growth",
                    "debt_to_equity", "operating_margin"]:
            result[col] = pd.to_numeric(result[col], errors="coerce")
        result["per"] = result["per"].round(1)
        result["pbr"] = result["pbr"].round(2)

        # PEG = PER ÷ (eps_growth × 100)
        def _peg(r):
            try:
                if pd.isna(r["per"]) or pd.isna(r["eps_growth"]):
                    return None   # 결측은 '이익 감소(50)'로 오표기하지 않고 빈칸 처리
                pe = float(r["per"])
                eg = float(r["eps_growth"])
                if pe > 0:
                    if eg > 0:
                        return round(pe / (eg * 100), 2)   # 정상 PEG
                    else:
                        return PEG_CAP                      # 이익 감소 → 최하
            except Exception:
                pass
            return None   # PER 없음(적자) → 가중치 제외
        result["peg"] = result.apply(_peg, axis=1)
        print(f"  PEG: {result['peg'].notna().sum()}건  "
              f"ROE: {result['roe'].notna().sum()}건  "
              f"RevGrowth: {result['revenue_growth'].notna().sum()}건  "
              f"OpMargin: {result['operating_margin'].notna().sum()}건")
    else:
        for col in ["per", "pbr", "peg", "roe", "revenue_growth",
                    "debt_to_equity", "operating_margin"]:
            result[col] = None
    result["tier"]         = result["current_rank"].apply(assign_tier)
    result["시총_성장률"]     = result["시총_성장률"].round(2)
    result["시총_가속도"]     = result["시총_가속도"].round(2)
    result["거래대금_증가율"]  = result["거래대금_증가율"].round(2)
    result["거래대금_가속도"]  = result["거래대금_가속도"].round(2)

    # 게이트(검증불가 팩터로 후보 필터) → 종합 점수 계산 → 정렬
    result, gated = apply_gates(result)
    if gated:
        print(f"  게이트 제외(ROE<{GATE_ROE_MIN:.0f} 또는 D/E>{GATE_DE_MAX:.0f}): "
              f"{len(gated)}종목 {gated[:20]}")
    result = calc_composite_score(result)
    result = result.sort_values("종합점수", ascending=False).reset_index(drop=True)

    # 안전(참고지표) 컬럼 추가 — 점수에 합산하지 않음(IC상 안전↔저수익)
    result = compute_safety_columns(result)

    # 출력
    print(f"\n{'='*110}")
    print("  결과 Top 20 (종합점수 기준)")
    _ws = sum(COMPOSITE_WEIGHTS.values()) or 1.0
    print(f"  가중치(IC검증 기반·실효비중): "
          f"매출가속도 {COMPOSITE_WEIGHTS.get('rev_acceleration',0)/_ws*100:.0f}%  "
          f"매출성장TTM {COMPOSITE_WEIGHTS.get('revenue_growth',0)/_ws*100:.0f}%  "
          f"매출일관성 {COMPOSITE_WEIGHTS.get('rev_consistency',0)/_ws*100:.0f}%  "
          f"(잔여 모멘텀 최소화, PEG/ROE/부채=게이트)")
    print(f"{'='*110}")
    print(f"{'순위':>4}  {'종목':>8}  {'티어':<18}  {'종합점수':>8}  "
          f"{'안전도':>6}  {'변동성%':>7}  {'사분면':<11}  "
          f"{'PER':>6}  {'PEG':>7}")
    print("-" * 110)
    for i, row in result.head(20).iterrows():
        per_s = f"{row['per']:>6.1f}" if pd.notna(row.get("per")) else f"{'N/A':>6}"
        peg_s = f"{row['peg']:>7.2f}" if pd.notna(row.get("peg")) else f"{'N/A':>7}"
        saf_s = f"{row['안전도']:>6.1f}" if pd.notna(row.get("안전도")) else f"{'N/A':>6}"
        vol_s = f"{row['변동성']:>7.1f}" if pd.notna(row.get("변동성")) else f"{'N/A':>7}"
        quad  = row.get("사분면", "") or ""
        print(f"{i+1:>4}  {row['code']:>8}  {row['tier']:<18}  "
              f"{row['종합점수']:>8.2f}  "
              f"{saf_s}  {vol_s}  {quad:<11}  "
              f"{per_s}  {peg_s}")

    # 상세 시트 정렬용: 종목별 점수(성장률)를 detail 에 붙임
    f1_detail = f1_detail.merge(
        f1_score[["code", "시총_성장률"]], on="code", how="left")
    f2_detail = f2_detail.merge(
        f2_score[["code", "거래대금_증가율"]], on="code", how="left")

    # 시총 Top 추이 데이터 (차트 탭용)
    trend_df, ref_dates = compute_trend_data(market, TREND_TOP_N)
    if trend_df is not None:
        print(f"  시총 Top{len(trend_df)} 추이: 5년전 {ref_dates.get('5년')} ~ 현재 {ref_dates.get('현재')}")

    # Excel 저장
    today      = date.today().strftime("%Y%m%d")
    out_path   = f"/data/frame/factor_result_{market}_{today}.xlsx"
    latest     = f"/data/frame/factor_result_{market}_latest.xlsx"
    save_excel(result, f1_detail, f2_detail, market, years, out_path, trend_df, ref_dates, name_map)
    save_excel(result, f1_detail, f2_detail, market, years, latest, trend_df, ref_dates, name_map)
    cleanup_old_results(market, out_path, latest)

    return result, out_path


def cleanup_old_results(market, keep_dated_path, keep_latest_path):
    """이번에 새로 만든 dated/latest 파일만 남기고, 같은 market의 이전 factor_result_*.xlsx는 삭제."""
    pattern = f"/data/frame/factor_result_{market}_*.xlsx"
    keep = {keep_dated_path, keep_latest_path}
    for f in glob.glob(pattern):
        if f not in keep:
            os.remove(f)
            print(f"  이전 결과 삭제: {f}")


# ── ETL 데이터 최신화 (--refresh) ─────────────────────────────────────────────
# run_factor_analysis 전에 ETL 스크립트를 순서대로 실행해 DB를 최신 상태로 만든다.
#   1) run_etl.py --us      : daily_price_us / daily_marcap_us (가격·시총, 증분)
#   2) etl_quarterly_sec.py : quarterly_financials_us (SEC 분기 매출)
#   3) etl_fundamental_us.py: daily_fundamental_us (PER/PBR/ROE/EPS/매출성장/영업이익률/부채비율, 오늘값)
# 한 단계가 실패해도 경고만 출력하고 다음 단계 및 스크리닝은 계속 진행한다.
BASE_DIR = "/data/frame"

ETL_STEPS = {
    "us": [
        ("가격·시총 (run_etl.py --us)",         [BASE_DIR + "/run_etl.py", "--us"]),
        ("SEC 분기매출 (etl_quarterly_sec.py)",  [BASE_DIR + "/etl_quarterly_sec.py"]),
        ("펀더멘털 (etl_fundamental_us.py)",     [BASE_DIR + "/etl_fundamental_us.py"]),
    ],
    "kr": [
        ("가격·시총 (run_etl.py --kr)",         [BASE_DIR + "/run_etl.py", "--kr"]),
    ],
}


def run_etl_refresh(market="us"):
    """ETL 스크립트를 순서대로 실행해 DB 데이터를 최신화한다 (실패해도 계속)."""
    steps = ETL_STEPS.get(market, [])
    if not steps:
        print(f"  ⚠️  '{market}' 시장은 --refresh 미지원 — ETL 건너뜀")
        return

    print("=" * 60)
    print(f"  [REFRESH] ETL 데이터 최신화 시작 — {market.upper()}")
    print("=" * 60)
    t_all = time.time()
    for label, script in steps:
        cmd = [sys.executable] + script
        print(f"\n▶ {label}")
        print(f"  $ {' '.join(cmd)}", flush=True)
        t0 = time.time()
        try:
            rc = subprocess.call(cmd, cwd=BASE_DIR)
        except Exception as e:
            print(f"  ⚠️  실행 예외: {e!r} — 다음 단계 계속")
            continue
        dt = time.time() - t0
        if rc == 0:
            print(f"  ✅ 완료 ({dt:.0f}s)")
        else:
            print(f"  ⚠️  비정상 종료 (exit {rc}, {dt:.0f}s) — 다음 단계 계속")
    print(f"\n[REFRESH] 전체 ETL 종료 ({time.time() - t_all:.0f}s) — 스크리닝 시작\n")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Factor 분석 (시가총액 & 거래대금)")
    parser.add_argument("--kr",    action="store_true", help="국장 분석")
    parser.add_argument("--years", type=int, default=4,   help="분석 기간 (년)")
    parser.add_argument("--top",   type=int, default=100, help="시총 상위 N개")
    parser.add_argument("--refresh", action="store_true",
                        help="스크리닝 전에 ETL을 실행해 데이터를 최신화")
    args = parser.parse_args()

    market = "kr" if args.kr else "us"
    if args.refresh:
        run_etl_refresh(market)
    result, out_path = run_factor_analysis(market=market, years=args.years, top_n=args.top)
    print(f"\n완료: {out_path}")
